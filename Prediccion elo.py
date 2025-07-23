import pandas as pd
import numpy as np
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_squared_error, r2_score
from scipy.stats import kendalltau
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

# Parámetros
BASE_RATING = 1500
K           = 32
ELO_RANGE   = 600
# regresor no lineal
model_cls   = GradientBoostingRegressor
model_args  = {
    "n_estimators": 200,
    "learning_rate": 0.1,
    "max_depth": 3,
    "random_state": 42
}

# Rutas
ruta_excel = "Sponsorship Prioritisation Form (respuestas).xlsx"
ruta_pairs = "Comparaciones.xlsx"

# 1) Carga datos
df_resp  = pd.read_excel(ruta_excel, sheet_name="Respuestas")[["Name","Urgency ranking"]]
df_pairs = pd.read_excel(ruta_pairs)[["Name A","Name B","Winner"]]
df_punt  = pd.read_excel(ruta_excel, sheet_name="Puntuaciones")

# 2) Elo transductivo
max_rank = df_resp["Urgency ranking"].max()
scale    = ELO_RANGE / (max_rank - 1)
df_resp["EloInitial"] = BASE_RATING + (max_rank - df_resp["Urgency ranking"]) * scale

ratings = dict(zip(df_resp["Name"], df_resp["EloInitial"]))
for _, r in df_pairs.iterrows():
    A, B = r["Name A"], r["Name B"]
    RA, RB = ratings[A], ratings[B]
    EA = 1/(1+10**((RB-RA)/400))
    EB = 1-EA
    SA, SB = (1.0,0.0) if r["Winner"]==A else (0.0,1.0)
    ratings[A] = RA + K*(SA-EA)
    ratings[B] = RB + K*(SB-EB)

df_elo = pd.DataFrame.from_dict(ratings, orient="index", columns=["EloAdjusted"])
df_elo.index.name = "Name"
df_train = df_resp.merge(df_elo, on="Name")

# 2.b) Compute EloAdjustedRank **NEW**
df_train["EloAdjustedRank"] = df_train["EloAdjusted"] \
    .rank(ascending=False, method="dense") \
    .astype(int)

# 3) Preparo X, y
feature_cols = [c for c in df_punt.columns if c.endswith(".p")] + ["EloInitial"]
df_feats     = df_punt.set_index("Name")[ [c for c in feature_cols if c!="EloInitial"] ]
df_model     = df_train.set_index("Name")[["EloAdjusted","EloInitial"]].join(df_feats, how="inner")

X = df_model[feature_cols].values
y = df_model["EloAdjusted"].values

# 4) Normalizo
scaler = StandardScaler()
X_norm = scaler.fit_transform(X)

# 5) Entreno el regresor no lineal
reg = model_cls(**model_args)
reg.fit(X_norm, y)

# 6) Evalúo en entrenamiento
y_pred = reg.predict(X_norm)
#print("MSE_train:", mean_squared_error(y,y_pred))
#print("R2_train:", r2_score(y,y_pred))
#print("Kendall τ:", kendalltau(y,y_pred)[0])

# 7) Predicción inductiva para todos los niños
df_all = df_feats.copy()
df_all["EloInitial"] = df_all.index.map(df_resp.set_index("Name")["EloInitial"])
X_all = df_all[feature_cols].values
X_all_norm = scaler.transform(X_all)
df_all["EloPred"]    = reg.predict(X_all_norm)
df_all["Urgency"] = df_all["EloPred"].rank(ascending=False, method="dense").astype(int)

# 9) Construir df de salida con Name, Project y Urgency, ordenado
df_out = df_all.reset_index().rename(columns={"index":"Name"})[["Name","Urgency"]]

# Traer “Project” de la hoja Respuestas y unirlo una sola vez
df_proj = pd.read_excel(ruta_excel, sheet_name="Respuestas")[["Name","Project"]]
df_out  = df_out.merge(df_proj, on="Name", how="left")

# Ordenar por Urgency ascendente (1 primero) y resetear índice
df_out = df_out.sort_values("Urgency", ascending=True).reset_index(drop=True)

# Ahora reordenamos columnas exactamente en el orden que queremos pintar
df_out = df_out[["Name","Project","Urgency"]]

# 10) Guardar en nueva hoja sin tocar la original
wb = load_workbook(ruta_excel)
if "Urgency ranking" in wb.sheetnames:
    wb.remove(wb["Urgency ranking"])
ws = wb.create_sheet("Urgency ranking")

# Encabezados (coinciden con df_out.columns)
for ci, col in enumerate(df_out.columns, start=1):
    ws.cell(row=1, column=ci, value=col)

# Datos
for ri, row in enumerate(df_out.itertuples(index=False), start=2):
    for ci, val in enumerate(row, start=1):
        ws.cell(row=ri, column=ci, value=val)

# 11) Aplicar gradiente de color de rojo a amarillo suave en la columna C
from openpyxl.formatting.rule import ColorScaleRule

n_rows = len(df_out) + 1  # incluye encabezado
rule = ColorScaleRule(
    start_type='num', start_value=1,       start_color='FF0000',
    end_type='num',   end_value=n_rows,    end_color='FFFFE0'
)
# Aplica al rango C2:C{n_rows}
ws.conditional_formatting.add(f"C2:C{n_rows}", rule)

wb.save(ruta_excel)

print("✅ Hoja 'Urgency ranking' actualizada con:")

