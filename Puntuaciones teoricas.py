import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

ruta_excel = "Sponsorship Prioritisation Form (respuestas).xlsx"
ruta_pairs = "Comparaciones.xlsx"

try:
    #Leer archivo y mostrar 5 primeras filas
    df = pd.read_excel(ruta_excel)
    print("Archivo leído con éxito: ")
    print(df.head())

    #Obtener numero de filas totales
    num_filas = len(df)
    print(f"El archivo tiene {num_filas} líneas (sin incluir el encabezado).")
    
    #Calcular edades a partir de las fechas de nacimiento
    current_year=datetime.now().year
    df['Age']=df['Date of birth'].apply(lambda x:current_year-x.year if pd.notnull(x) else None)

    #Clasificar y puntuar las edades
    def calcular_puntuacion_edad(edad):
        if edad is None:
            return None
        elif edad < 3:
            return 0
        elif 3 <= edad < 6:
            return 2
        elif 6 <= edad < 9:
            return 6
        elif 9 <= edad < 12:
            return 10
        elif 12 <= edad < 15:
            return 8
        elif 15 <= edad < 18:
            return 5
        else:
            return 4

    df['D'] = df['Age'].apply(calcular_puntuacion_edad)

    # Clasificar puntuación basada en si va al colegio (sale en formato hora, 
    # por lo que no se puede hacer directo)
    def calcular_puntuacion_escuela(valor):
        if valor == 'Yes':
            return 0
        elif valor == 'No':
            return 20
        else:
            return None

    # Puntuación para 'Percentage of kids going school'
    def calcular_puntuacion_porcentaje(row):
        if row['Number of underage in the family '] == 0:  # Evitar división por cero
            return 0
        porcentaje = (row['Number of kids going school'] / row['Number of underage in the family ']) * 100
        if 0 <= porcentaje <= 25:
            return 10
        elif 25 < porcentaje <= 50:
            return 6
        elif 50 < porcentaje <= 75:
            return 3
        elif 75 < porcentaje <= 100:
            return 0
        else:
            return None

    # Definir función para calcular la puntuación del porcentaje de colchones por miembro de la familia
    def calcular_puntuacion_colchones(row):
        total_miembros = row['Number of adults in the family '] + row['Number of underage in the family ']
        if total_miembros == 0: # Evitar división por cero
            return None
        porcentaje_colchones = (row['Number of mattresses per family'] / total_miembros) * 100
        if 0 <= porcentaje_colchones <= 25:
            return 5
        elif 25 < porcentaje_colchones <= 50:
            return 4
        elif 50 < porcentaje_colchones <= 75:
            return 2
        elif 75 < porcentaje_colchones <= 100:
            return 1
        elif 100 < porcentaje_colchones:
            return 0
        else:
            return None

    # Nueva función para calcular la puntuación del porcentaje de mosquiteras
    def calcular_puntuacion_mosquiteras(row):
        total_miembros = row['Number of adults in the family '] + row['Number of underage in the family ']
        if total_miembros == 0:  # Evitar división por cero
            return None
        porcentaje_mosquiteras = (row['Number of mosquito nets per family'] / total_miembros) * 100
        if 0 <= porcentaje_mosquiteras <= 25:
            return 5
        elif 25 < porcentaje_mosquiteras <= 50:
            return 4
        elif 50 < porcentaje_mosquiteras <= 75:
            return 2
        elif 75 < porcentaje_mosquiteras <= 100:
            return 1
        elif 100 < porcentaje_mosquiteras:
            return 0
        else:   
            return None

    # Calcular cuanto llevan en el proyecto (en meses)
    current_date = pd.Timestamp(datetime.now())
    df['MonthsInProject'] = ((current_date - df['When did you start being part of the project? ']).dt.days / 30).round()
    # Clasificar los rangos
    df['ProjectTime.p'] = df['MonthsInProject'].apply(lambda months: 0 if months <= 6 
                                                                else 3 if 6 < months <= 12 
                                                                else 7 if 12 < months <= 24 
                                                                else 11 if 24 < months <= 36 
                                                                else 15 if months > 36 
                                                                else None)

    #Clasificar y puntuar las respuestas obtenidas
    puntuaciones_df=pd.DataFrame({
        'Name':df['Name'],
        'Sex.p' : df['Sex'].apply(lambda x:0 if x=='Male' 
                                        else 10 if x=='Female' 
                                        else None),
        'Age.p': df['D'],
        'School.p' : df['Is the kid currently attending school?'].apply(calcular_puntuacion_escuela),
        'FinancialSupport.p': df['Who pays the school fees?'].apply(lambda x: 5 if x == "Himself/herself" 
                                                                            else 3 if x == "Family member" 
                                                                            else 2 if x == "Parent" 
                                                                            else 0 if x in ["Other sponsor", "Cooperating NGO", "Do not apply"] 
                                                                            else None),
        'ClassStart.p': df['Which class do you have to start?'].apply(lambda x: 5 if x in ["Nursery", "Baby Class", "Middle Class", "Top Class"] 
                                                                            else 12 if x in ["P1", "P2", "P3", "S1", "S2"] 
                                                                            else 15 if x in ["P4", "P5", "P6", "P7"] 
                                                                            else 10 if x in ["S3", "S4"] 
                                                                            else 7 if x in ["S5", "S6"] 
                                                                            else 5 if x == "Higher studies" 
                                                                            else None),
        'Grade.p': df['Average grade of the last course'].apply(lambda x: 0 if x == "0-50" 
                                                                            else 6 if x == "50-70" 
                                                                            else 9 if x == "70-80" 
                                                                            else 12 if x == "80-90" 
                                                                            else 15 if x == "90-100" 
                                                                            else 0 if x == "Do not apply" 
                                                                            else None),
        'FamilyHead.p': df['Who is the head of the family? '].apply(lambda x: 0 if x == "Mother and father" 
                                                                            else 6 if x in ["Single parent", "Other relative (adult)"] 
                                                                            else 8 if x == "Other relative (underage)" 
                                                                            else 9 if x == "Orphan" 
                                                                            else 10 if x == "Alone" 
                                                                            else None),
        'Adults.p': df['Number of adults in the family '].apply(lambda x: 5 if x == 0 
                                                                        else 3 if x == 1
                                                                        else 1 if x == 2 
                                                                        else 0 if x >= 3
                                                                        else None), 
        'Underage.p': df['Number of underage in the family '].apply(lambda x: 0 if x == 0 
                                                            else 1 if 1 <= x <= 2 
                                                            else 4 if 3 <= x <= 5 
                                                            else 8 if 6 <= x <= 8 
                                                            else 10 if x >= 9 
                                                            else None),
        'SchoolPercentage.p': df.apply(calcular_puntuacion_porcentaje, axis=1),
        'FamilyRelationship.p': df['How is the family relationship?'].apply(lambda x: 0 if x == "Very good" 
                                                                                    else 3 if x == "Good" 
                                                                                    else 7 if x == "Bad" 
                                                                                    else 10 if x == "Very bad" 
                                                                                    else None),
        'JobType.p': df['Type of job of the head of family'].apply(lambda x: 0 if x == "Self-employed" 
                                                                            else 4 if x == "Employee" 
                                                                            else 8 if x == "Part-time earning" 
                                                                            else 10 if x == "Not working" 
                                                                            else None), 
        'Income.p': df['Total family income per month '].apply(lambda x: 20 if x < 50000
                                                                    else 16 if 50000 <= x < 100000
                                                                    else 12 if 100000 <= x < 150000
                                                                    else 8 if 150000 <= x < 200000
                                                                    else 4 if x >= 200000
                                                                    else None),
        'Land.p': df['Is the land owned or rented? '].apply(lambda x: 0 if x == 'Owned' 
                                                                else 1 if x == 'Land caregiver' 
                                                                else 2 if x == 'Rented' 
                                                                else None),
        'House.p': df['Type of house'].apply(lambda x: 0 if x == 'Permanent' 
                                                    else 2 if x == 'Semi-permanent' 
                                                    else None),
        'HouseConditions.p': df['Condition of the house '].apply(lambda x: 0 if x == 'Very good' 
                                                                        else 2 if x == 'Good' 
                                                                        else 7 if x == 'Bad' 
                                                                        else 10 if x == 'Very bad' 
                                                                        else None),
        'Garden.p': df['Is there a garden?'].apply(lambda x: 0 if x == 'Yes' 
                                                        else 2 if x == 'No' 
                                                        else None),
        'Animals.p': df['Do you have animals?'].apply(lambda x: 0 if x in ['Cows', 'Pigs', 'Goats'] 
                                                            else 1 if x in ['Birds', 'Rabbits', 'Others']  
                                                            else 2 if x == 'No' 
                                                            else None),
        'LightSource.p': df['Light source '].apply(lambda x: 0 if x == 'Electricity' 
                                                        else 2 if x == 'Solar' 
                                                        else 3 if x == 'Torch' 
                                                        else 4 if x == 'Candles' 
                                                        else 5 if x == 'Natural light' 
                                                        else None),
        'WaterSource.p': df['Water source '].apply(lambda x: 5 if x == 'River' 
                                                        else 3 if x == 'Spring well' 
                                                        else 2 if x == 'Boraholds' 
                                                        else 0 if x == 'Commercial water' 
                                                        else None),
        'MattressPercentage.p': df.apply(calcular_puntuacion_colchones, axis=1),
        'MosquitoNetPercentage.p': df.apply(calcular_puntuacion_mosquiteras, axis=1),
        'ClosedKitchen.p': df['Do they have a closed kitchen? '].apply(lambda x: 0 if x == 'Yes' else 1 if x == 'No' else None),
        'MealsPerDay.p': df['How many meals do you eat per day? '].apply(lambda x: 15 if x == 1 
                                                                                else 7 if x == 2 
                                                                                else 0 if x == '3 or more' 
                                                                                else None),
        'Illness.p': df['Does the kid have any illness?'].apply(lambda x: 0 if x == 'No'  
                                                                else 2 if x in ['Diabetes', 'Hipertension', 'Other']
                                                                else 4 if x == 'Asthma' 
                                                                else 6 if x == 'Sickle cells'
                                                                else 8 if x == 'HIV' 
                                                                else 10 if x in ['Cancer', 'Blind/Deaf', 'Reduced mobility', 'Epilepsy', "Down's Syndrome", 'Mental illness'] 
                                                                else None),
        'Treatment.p': df.apply(lambda row: 0 if row['Does the kid have any illness?'] == 'No' 
                                        else 0 if row['Is the kid on treatment?'] == 'Yes' 
                                        else 15 if row['Is the kid on treatment?'] == 'No' 
                                        else None, axis=1),
        'HeadFamilyIllness.p': df['Does the head of the family have any illness? '].apply(lambda x: 0 if x == 'No'  
                                                                else 1 if x in ['Diabetes', 'Hipertension', 'Other']
                                                                else 2 if x == 'Asthma' 
                                                                else 3 if x == 'Sickle cells'
                                                                else 4 if x == 'HIV' 
                                                                else 5 if x in ['Cancer', 'Blind/Deaf', 'Reduced mobility', 'Epilepsy', "Down's Syndrome", 'Mental illness'] 
                                                                else None),
        'HeadFamilyTreatment.p': df.apply(lambda row: 0 if row['Does the head of the family have any illness? '] == 'No' 
                                                else 0 if row['Is the head of the family on treatment?'] == 'Yes' 
                                                else 15 if row['Is the head of the family on treatment?'] == 'No' 
                                                else None, axis=1),
        'Addiction.p': df['Are there any members with addiction problems? '].apply(lambda x: 0 if x == 'No' 
                                                                                        else 10 if x in ['Drugs', 'Alcohol'] 
                                                                                        else 8 if x in ['Gambling', 'Gumbling'] 
                                                                                        else None),
        'Latrine.p': df['How is the latrine like? '].apply(lambda x: 5 if x == 'No latrine' 
                                                                else 3 if x == 'Shared latrine' 
                                                                else 1 if x == 'Own latrine' 
                                                                else 0 if x == 'Self contained latrine' 
                                                                else None),
        'Shower.p': df['How many times do you shower per week? '].apply(lambda x: 2 if x == 0 
                                                                        else 1 if 1 <= x <= 2 
                                                                        else 0 if x >= 3 
                                                                        else None),
        'Privacy.p': df['Is there privacy in the shower? '].apply(lambda x: 0 if x == 'Yes' 
                                                                        else 2 if x == 'No' 
                                                                        else None),
        'Pads.p': df.apply(lambda row: 0 if row['Sex'] == 'Male' 
                                else 0 if row['What type of pads do you use? '] in ['Do not apply', 'Disposable pads'] 
                                else 5 if row['What type of pads do you use? '] == 'Reusable pads' 
                                else 10 if row['What type of pads do you use? '] == 'Clothes/Towels' 
                                else None, axis=1),
        'ProjectTime.p': df['ProjectTime.p'],
        'SponsoredKids.p': df.apply(lambda row: 0 if row['Number of underage in the family '] == 0 
                                            else 5 if 0 <= (row['How many kids are being sponsored in this family? '] / row['Number of underage in the family ']) * 100 <= 25
                                            else 2 if 25 < (row['How many kids are being sponsored in this family? '] / row['Number of underage in the family ']) * 100 <= 50
                                            else 1 if 50 < (row['How many kids are being sponsored in this family? '] / row['Number of underage in the family ']) * 100 <= 75
                                            else 0 if  75 < (row['How many kids are being sponsored in this family? '] / row['Number of underage in the family ']) * 100 <= 100
                                            else None, axis=1),
        'PreviouslySponsored.p': df['Have you previously been sponsored by Cooperating NGO? '].apply(lambda x: 20 if x == "Yes" 
                                                                                                            else 0 if x == "No" 
                                                                                                            else None)
    })

    #Cargar el archivo Excel
    book=load_workbook(ruta_excel)

    #Crear nuevo hoja para las puntuaciones
    if "Puntuaciones" not in book.sheetnames:
        book.create_sheet("Puntuaciones")

    #Obtener la hoja creada
    sheet=book["Puntuaciones"]

    #Limpiar la hoja para evitar datos antiguos
    for row in sheet.iter_rows():
        for cell in row:
            cell.value=None

    #Escribir los encabezados
    for c_idx, col_name in enumerate(puntuaciones_df.columns, start=1):
        sheet.cell(row=1, column=c_idx, value=col_name)

    from openpyxl.styles import numbers

    # Escribir los datos en la nueva hoja sin sobrescribir
    for r_idx, row in enumerate(puntuaciones_df.values, start=2):  # Fila de encabezado ya ocupada
        for c_idx, value in enumerate(row, start=1):  # Iterar por columnas
            cell = sheet.cell(row=r_idx, column=c_idx)
            if isinstance(value, (int, float)):  # Si es número
                cell.value = value
                cell.number_format = numbers.FORMAT_NUMBER  # Formato explícito para números
            else:  # Si es texto (o None)
                cell.value = value  # Guardar como texto

    #Guardar el archivo excel
    book.save(ruta_excel)

    #Mostrar los resultados de la hoja 'Puntuaciones'
    df=pd.read_excel(ruta_excel,sheet_name="Puntuaciones")
    df["TOTAL"] = df.select_dtypes(include="number").sum(axis=1)
    print("Contenido de la hoja de 'Puntuaciones': ")
    print(df)
    with pd.ExcelWriter(ruta_excel, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name="Puntuaciones", index=False)

    # Mostrar cuántos valores nulos hay en cada columna
    print("Valores nulos por columna:")
    print(df.isnull().sum())

    print(f"Datos calculados se han añadido correctamente en la hoja 'Puntuaciones'.")

    #Crear ranking de priorización según la urgencia
    df_puntuaciones = pd.read_excel(ruta_excel, sheet_name="Puntuaciones")
    df_respuestas = pd.read_excel(ruta_excel, sheet_name="Respuestas")
   
   # 1) Calcula el ranking en df_puntuaciones
    df_puntuaciones["Urgency ranking"] = (
        df_puntuaciones["TOTAL"]
          .rank(ascending=False, method="dense")
          .astype(int)
    )
    # Mapa Name → ranking
    ranking_map = df_puntuaciones.set_index("Name")["Urgency ranking"]

    # 2) Elimina cualquier columna antigua que empiece por 'Urgency ranking'
    old_cols = [c for c in df_respuestas.columns if c.startswith("Urgency ranking")]
    if old_cols:
        df_respuestas.drop(columns=old_cols, inplace=True)

    # 3) Asigna la nueva columna (la crea o la sobreescribe)
    df_respuestas["Urgency ranking"] = df_respuestas["Name"].map(ranking_map)

    # Sort the DataFrame so rank 1 in order
    df_respuestas.sort_values(by="Urgency ranking", ascending=True, inplace=True)

    # 4) Escribe de nuevo la hoja 'Respuestas', reemplazando la existente
    with pd.ExcelWriter(ruta_excel, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_respuestas.to_excel(writer, sheet_name="Respuestas", index=False)

    print("✅ ‘Urgency ranking’ ha sido creado o actualizado sin duplicados.")

except FileNotFoundError:
    print(f"El archivo '{ruta_excel}' no se encuentra.")    

except Exception as e:
    print(f"Ocurrió un error al leer el archivo: {e}")
