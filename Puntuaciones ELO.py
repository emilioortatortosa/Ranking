import pandas as pd
from openpyxl import load_workbook

# Parámetros Elo
BASE_RATING = 1500
K = 32

# 1) Rutas a tus archivos
ruta_excel = "Sponsorship Prioritisation Form (respuestas).xlsx"
ruta_pairs = "Comparaciones.xlsx"

# 2) Cargo tus datos SIN tocar la hoja original
df = pd.read_excel(ruta_excel, sheet_name="Respuestas")[["Name", "Urgency ranking"]]
pairs = pd.read_excel(ruta_pairs)[["Name A", "Name B", "Winner"]]

# 3) Inicializo Elo “teórico” a partir del ordinal
max_rank = df["Urgency ranking"].max()
scale    = 600 / (max_rank - 1)
df["EloInitial"] = BASE_RATING + (max_rank - df["Urgency ranking"]) * scale

# 4) Aplico las 54 comparaciones para ajustar los ratings
ratings = dict(zip(df["Name"], df["EloInitial"]))
for _, row in pairs.iterrows():
    A, B = row["Name A"], row["Name B"]
    RA, RB = ratings[A], ratings[B]
    EA = 1 / (1 + 10 ** ((RB - RA) / 400))
    EB = 1 - EA
    if row["Winner"] == A:
        SA, SB = 1.0, 0.0
    else:
        SA, SB = 0.0, 1.0
    ratings[A] = RA + K * (SA - EA)
    ratings[B] = RB + K * (SB - EB)

# 5) Monto un DataFrame con los ratings ajustados
df_ratings = pd.DataFrame.from_dict(ratings, orient="index", columns=["EloAdjusted"])
df_ratings.index.name = "Name"
df = df.merge(df_ratings, on="Name")

# 6) Calculo el ranking final Elo (1 = más urgente)
df["HybridRankElo"] = df["EloAdjusted"].rank(ascending=False, method="dense").astype(int)

# 7) Calculo e imprimo Kendall Tau
tau = df["Urgency ranking"].corr(df["HybridRankElo"], method="kendall")
print(f"Kendall Tau (Urgency ranking vs HybridRankElo): {tau:.3f}")

# 8) Preparo el DataFrame de salida con Elo inicial, Elo ajustado y HybridRank
df_out = df[[
    "Name",
    "Urgency ranking",
    "EloInitial",
    "EloAdjusted",
    "HybridRankElo"
]]

# 9) Escribo en hoja nueva 'Respuestas_Elo' sin tocar la original
wb = load_workbook(ruta_excel)
if "Respuestas_Elo" in wb.sheetnames:
    wb.remove(wb["Respuestas_Elo"])
ws = wb.create_sheet("Respuestas_Elo")

# 10) Encabezados
for col_idx, col in enumerate(df_out.columns, start=1):
    ws.cell(row=1, column=col_idx, value=col)

# 11) Filas de datos
for row_idx, row in enumerate(df_out.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 12) Escribo el Tau dos filas más abajo
ws.cell(row=len(df_out) + 2, column=1, value="Kendall Tau:")
ws.cell(row=len(df_out) + 2, column=2, value=round(tau, 3))

# 13) Guardar sin modificar 'Respuestas'
wb.save(ruta_excel)
print("✅ Hoja 'Respuestas_Elo' creada con EloInitial, EloAdjusted, HybridRankElo y Tau.")
